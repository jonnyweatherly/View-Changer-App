"""
Database seeding script for the View Changer App.

Reads the Excel spreadsheet and populates the database with:
- Patient Columns (master list of available columns)
- View Columns (current view configuration)

Run locally:
    python scripts/seed_db.py
    
To re-seed (clear and reload):
    python scripts/seed_db.py --reset
"""

import os
import sys

# Add parent directory to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from models import SessionLocal, PatientColumn, ViewColumn, Base, engine
import openpyxl


def seed_database(reset: bool = False):
    """Seed the database from the Excel spreadsheet."""
    
    # Find the Excel file
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_dir = os.path.dirname(script_dir)
    excel_path = os.path.join(project_dir, "View Columns 1.xlsx")
    
    if not os.path.exists(excel_path):
        # Try data folder
        excel_path = os.path.join(project_dir, "data", "seed.xlsx")
    
    if not os.path.exists(excel_path):
        print(f"Error: Could not find Excel file at {excel_path}")
        return False
    
    print(f"Loading spreadsheet: {excel_path}")
    
    # Reset database if requested
    if reset:
        print("Resetting database...")
        Base.metadata.drop_all(bind=engine)
        Base.metadata.create_all(bind=engine)
    
    db = SessionLocal()
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path)
        
        # Seed Patient Columns
        print("\n--- Seeding Patient Columns ---")
        if "Patient Columns" in wb.sheetnames:
            ws = wb["Patient Columns"]
            count = 0
            for row in ws.iter_rows(min_row=2):  # Skip header
                values = [cell.value for cell in row]
                if values[0]:  # Has table name
                    # Check if already exists
                    existing = db.query(PatientColumn).filter(
                        PatientColumn.table_name == values[0],
                        PatientColumn.column_name == values[1]
                    ).first()
                    
                    if not existing:
                        col = PatientColumn(
                            table_name=values[0] or "",
                            column_name=values[1] or "",
                            data_type=values[2] or "varchar",
                            label=values[3] or values[1] or ""
                        )
                        db.add(col)
                        count += 1
            
            db.commit()
            print(f"Added {count} patient columns")
        else:
            print("Warning: 'Patient Columns' tab not found")
        
        # Seed View Columns
        print("\n--- Seeding View Columns ---")
        if "View Columns" in wb.sheetnames:
            ws = wb["View Columns"]
            count = 0
            for row in ws.iter_rows(min_row=2):  # Skip header
                values = [cell.value for cell in row]
                if values[0]:  # Has view name
                    # Check if already exists
                    existing = db.query(ViewColumn).filter(
                        ViewColumn.view_name == values[0],
                        ViewColumn.column_name == values[2]
                    ).first()
                    
                    if not existing:
                        # Parse width (remove 'px' if present for storage, but keep it)
                        width = str(values[6]) if values[6] else "100px"
                        if not width.endswith("px"):
                            width = width + "px"
                        
                        col = ViewColumn(
                            view_name=values[0] or "vwTitanium_WLAdmin",
                            table_name=values[1] if values[1] != "NULL" else None,
                            column_name=values[2] or "",
                            data_type=values[3] or "varchar",
                            label=values[4] or values[2] or "",
                            grid_order=int(values[5]) if values[5] is not None else 0,
                            grid_width=width,
                            is_right_aligned=bool(values[7]) if len(values) > 7 else False
                        )
                        db.add(col)
                        count += 1
            
            db.commit()
            print(f"Added {count} view columns")
        else:
            print("Warning: 'View Columns' tab not found")
        
        # Print summary
        patient_count = db.query(PatientColumn).count()
        view_count = db.query(ViewColumn).count()
        print(f"\n=== Database Summary ===")
        print(f"Total patient columns: {patient_count}")
        print(f"Total view columns: {view_count}")
        
        return True
        
    except Exception as e:
        print(f"Error seeding database: {e}")
        db.rollback()
        return False
    finally:
        db.close()


if __name__ == "__main__":
    reset = "--reset" in sys.argv
    success = seed_database(reset=reset)
    sys.exit(0 if success else 1)
